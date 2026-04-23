import json
from datetime import datetime
from typing import List, Optional

from fastapi import APIRouter, Depends, HTTPException, UploadFile
from sqlalchemy.orm import Session

from api.deps import get_db
from models.product import Product
from models.product_match import ProductMatch
from models.store import Store
from models.supplier_price import SupplierPrice
from services import embedding_service
from services.matching_service import get_candidates, get_supplier_similar, _score_str
from services.yam_client import get_client_for_store
from utils.normalizer import normalize_name

router = APIRouter(prefix="/matching", tags=["matching"])


@router.get("/stats")
def get_matching_stats(supplier: Optional[str] = None, db: Session = Depends(get_db)):
    q = db.query(ProductMatch)
    if supplier:
        q = q.filter(ProductMatch.supplier == supplier)
    all_matches = q.all()
    return {
        "pending":    sum(1 for m in all_matches if m.status == "pending"),
        "confirmed":  sum(1 for m in all_matches if m.status == "confirmed"),
        "stoplist":   sum(1 for m in all_matches if m.status == "stoplist"),
        "no_price":   sum(1 for m in all_matches if m.status == "no_price"),
        "auto_review": sum(1 for m in all_matches if m.status == "auto_review"),
    }


@router.get("")
def get_matching(status: Optional[str] = None, supplier: Optional[str] = None, db: Session = Depends(get_db)):
    q = db.query(ProductMatch)
    if status:
        q = q.filter(ProductMatch.status == status)
    if supplier:
        q = q.filter(ProductMatch.supplier == supplier)
    matches = q.order_by(ProductMatch.created_at.desc()).all()
    current_prices: set[tuple[str, str]] = {
        (sp.supplier, sp.normalized_name)
        for sp in db.query(SupplierPrice.supplier, SupplierPrice.normalized_name).all()
        if sp.supplier and sp.normalized_name
    }
    return [
        {
            "id": m.id, "supplier": m.supplier, "supplier_name": m.supplier_name,
            "supplier_normalized": m.supplier_normalized, "supplier_price": m.supplier_price,
            "sku": m.sku, "store_id": m.store_id, "product_name": m.product_name,
            "status": m.status, "match_type": m.match_type, "best_score": m.best_score,
            "created_at": m.created_at.isoformat() if m.created_at else None,
            "price_is_current": (m.supplier or "", m.supplier_normalized or "") in current_prices,
            "confirmed_at": m.confirmed_at.isoformat() if m.confirmed_at else None,
        }
        for m in matches
    ]


@router.post("/{match_id}/confirm")
def confirm_match(match_id: int, sku: str, store_id: int, db: Session = Depends(get_db)):
    m = db.query(ProductMatch).filter(ProductMatch.id == match_id).first()
    if not m:
        raise HTTPException(status_code=404, detail="Not found")
    product = db.query(Product).filter(Product.store_id == store_id, Product.sku == sku).first()
    m.sku = sku
    m.store_id = store_id
    m.product_name = product.name if product else sku
    m.status = "confirmed"
    m.match_type = "manual"
    m.auto_match = True
    db.commit()
    db.refresh(m)
    return {"id": m.id, "status": m.status, "sku": m.sku, "product_name": m.product_name}


@router.post("/{match_id}/keep-price")
def keep_old_price(match_id: int, price: Optional[float] = None, db: Session = Depends(get_db)):
    m = db.query(ProductMatch).filter(ProductMatch.id == match_id).first()
    if not m:
        raise HTTPException(status_code=404, detail="Not found")
    if price is not None:
        m.supplier_price = price
    m.status = "confirmed"
    db.commit()
    return {"id": m.id, "status": m.status, "supplier_price": m.supplier_price}


@router.post("/{match_id}/zero-stock")
def zero_stock_match(match_id: int, db: Session = Depends(get_db)):
    m = db.query(ProductMatch).filter(ProductMatch.id == match_id).first()
    if not m:
        raise HTTPException(status_code=404, detail="Not found")
    was_no_price = m.status == "no_price"
    if m.sku and m.store_id:
        other_confirmed = db.query(ProductMatch).filter(
            ProductMatch.sku == m.sku, ProductMatch.store_id == m.store_id,
            ProductMatch.status == "confirmed", ProductMatch.id != m.id,
        ).first()
        if not other_confirmed:
            store = db.query(Store).filter(Store.id == m.store_id).first()
            if store:
                try:
                    client, campaign_ids, _ = get_client_for_store(store.name, db=db)
                    for cid in campaign_ids:
                        client.update_stock(cid, m.sku, 0)
                except Exception:
                    pass
    m.status = "awaiting_price" if was_no_price else "stoplist"
    db.commit()
    return {"id": m.id, "status": m.status}


@router.post("/zero-all-no-price")
def zero_all_no_price(db: Session = Depends(get_db)):
    matches = db.query(ProductMatch).filter(ProductMatch.status == "no_price").all()
    zeroed = 0
    errors = 0
    no_price_sku_store = {(m.sku, m.store_id) for m in matches if m.sku and m.store_id}
    no_price_ids = {m.id for m in matches}
    protected: set[tuple] = set()
    if no_price_sku_store:
        for other in db.query(ProductMatch).filter(
            ProductMatch.status == "confirmed",
            ProductMatch.id.notin_(no_price_ids),
        ).all():
            if (other.sku, other.store_id) in no_price_sku_store:
                protected.add((other.sku, other.store_id))
    for m in matches:
        if m.sku and m.store_id and (m.sku, m.store_id) not in protected:
            store = db.query(Store).filter(Store.id == m.store_id).first()
            if store:
                try:
                    client, campaign_ids, _ = get_client_for_store(store.name, db=db)
                    for cid in campaign_ids:
                        client.update_stock(cid, m.sku, 0)
                except Exception:
                    errors += 1
        m.status = "awaiting_price"
        zeroed += 1
    db.commit()
    return {"zeroed": zeroed, "errors": errors}


@router.post("/{match_id}/restore-no-price")
def restore_no_price(match_id: int, db: Session = Depends(get_db)):
    m = db.query(ProductMatch).filter(ProductMatch.id == match_id).first()
    if not m:
        raise HTTPException(status_code=404, detail="Not found")
    m.status = "no_price"
    db.commit()
    return {"id": m.id, "status": m.status}


@router.post("/{match_id}/stoplist")
def stoplist_match(match_id: int, db: Session = Depends(get_db)):
    m = db.query(ProductMatch).filter(ProductMatch.id == match_id).first()
    if not m:
        raise HTTPException(status_code=404, detail="Not found")
    m.status = "stoplist"
    m.sku = None
    m.store_id = None
    m.product_name = None
    db.commit()
    return {"id": m.id, "status": m.status}


@router.post("/{match_id}/reset")
def reset_match(match_id: int, db: Session = Depends(get_db)):
    m = db.query(ProductMatch).filter(ProductMatch.id == match_id).first()
    if not m:
        raise HTTPException(status_code=404, detail="Not found")
    old_sku, old_store_id = m.sku, m.store_id
    if old_sku and old_store_id:
        other_confirmed = db.query(ProductMatch).filter(
            ProductMatch.sku == old_sku, ProductMatch.store_id == old_store_id,
            ProductMatch.status == "confirmed", ProductMatch.id != m.id,
        ).first()
        if not other_confirmed:
            store = db.query(Store).filter(Store.id == old_store_id).first()
            if store:
                try:
                    client, campaign_ids, _ = get_client_for_store(store.name, db=db)
                    for cid in campaign_ids:
                        client.update_stock(cid, old_sku, 0)
                except Exception:
                    pass
            product = db.query(Product).filter(Product.store_id == old_store_id, Product.sku == old_sku).first()
            if product:
                product.stock = 0
        m.blocked_sku = old_sku
        m.blocked_store_id = old_store_id
    m.auto_match = False
    m.status = "pending"
    m.sku = None
    m.store_id = None
    m.product_name = None
    m.match_type = None
    db.commit()
    return {"id": m.id, "status": m.status}


@router.post("/{match_id}/approve-auto")
def approve_auto_match(match_id: int, db: Session = Depends(get_db)):
    m = db.query(ProductMatch).filter(ProductMatch.id == match_id).first()
    if not m:
        raise HTTPException(status_code=404, detail="Not found")
    if m.status == "auto_review":
        m.status = "confirmed"
        m.confirmed_at = datetime.utcnow()
        db.commit()
    return {"id": m.id, "status": m.status, "confirmed_at": m.confirmed_at.isoformat() if m.confirmed_at else None}


@router.post("/{match_id}/reject-auto")
def reject_auto_match(match_id: int, db: Session = Depends(get_db)):
    m = db.query(ProductMatch).filter(ProductMatch.id == match_id).first()
    if not m:
        raise HTTPException(status_code=404, detail="Not found")
    if m.status == "auto_review":
        if m.sku and m.store_id:
            m.blocked_sku = m.sku
            m.blocked_store_id = m.store_id
        m.auto_match = False
        m.status = "pending"
        m.sku = None
        m.store_id = None
        m.product_name = None
        m.match_type = None
        db.commit()
    return {"id": m.id, "status": m.status}


@router.get("/{match_id}/supplier-similar")
def match_supplier_similar(match_id: int, db: Session = Depends(get_db)):
    m = db.query(ProductMatch).filter(ProductMatch.id == match_id).first()
    if not m:
        raise HTTPException(status_code=404, detail="Not found")
    return get_supplier_similar(m.supplier_normalized or "", m.supplier or "", db)


@router.get("/{match_id}/candidates")
def match_candidates(match_id: int, db: Session = Depends(get_db)):
    m = db.query(ProductMatch).filter(ProductMatch.id == match_id).first()
    if not m:
        raise HTTPException(status_code=404, detail="Not found")
    products = db.query(Product).filter(Product.enabled == True).all()  # noqa: E712
    return get_candidates(m.supplier_normalized, products, top_n=10)


@router.get("/export-pending")
def export_pending_matches(db: Session = Depends(get_db)):
    pending = db.query(ProductMatch).filter(ProductMatch.status == "pending").all()
    products = db.query(Product).filter(Product.enabled == True).all()  # noqa: E712
    return [
        {
            "supplier": m.supplier, "supplier_name": m.supplier_name,
            "supplier_price": m.supplier_price,
            "candidates": get_candidates(m.supplier_normalized, products, top_n=6),
        }
        for m in pending
    ]


@router.get("/export-unmatched-skus")
def export_unmatched_skus(db: Session = Depends(get_db)):
    matched_pairs = {
        (r.sku, r.store_id)
        for r in db.query(ProductMatch.sku, ProductMatch.store_id)
        .filter(ProductMatch.status.in_(["confirmed", "auto_review"]), ProductMatch.sku.isnot(None)).all()
    }
    products = db.query(Product).filter(Product.enabled == True).all()  # noqa: E712
    all_suppliers = db.query(SupplierPrice).all()
    stores = {s.id: (s.display_name or s.name) for s in db.query(Store).all()}
    result = []
    for p in products:
        if (p.sku, p.store_id) not in matched_pairs:
            sku_norm = normalize_name(p.sku)
            name_norm = p.normalized_name or normalize_name(p.name or "")
            scored = []
            for sp in all_suppliers:
                if not sp.normalized_name:
                    continue
                score = max(
                    _score_str(sku_norm, sp.normalized_name) if sku_norm else 0.0,
                    _score_str(name_norm, sp.normalized_name) if name_norm else 0.0,
                )
                if score >= 0.15:
                    scored.append({
                        "supplier": sp.supplier, "supplier_name": sp.name,
                        "supplier_price": sp.price, "score": round(score, 3),
                    })
            scored.sort(key=lambda x: -x["score"])
            result.append({
                "store": stores.get(p.store_id, str(p.store_id)),
                "sku": p.sku, "name": p.name, "price": p.price, "candidates": scored[:6],
            })
    return result


@router.post("/import-supplier-matches")
async def import_supplier_matches(file: UploadFile, db: Session = Depends(get_db)):
    import io
    import openpyxl
    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active
    headers = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
    stores_by_name = {(s.display_name or s.name): s.id for s in db.query(Store).all()}
    stores_by_name.update({"ЯМ16": 1, "ЯМ21": 2})
    updated = 0
    errors = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        data = dict(zip(headers, row))
        choice_raw = data.get("Выбор (1-6)")
        if not choice_raw:
            continue
        try:
            n = int(str(choice_raw).strip())
        except ValueError:
            errors.append(f"Неверное значение Выбор: {choice_raw}"); continue
        if n < 1 or n > 6:
            errors.append(f"Выбор должен быть 1-6, получено: {n}"); continue
        supplier = str(data.get("Поставщик") or "").strip()
        supplier_name = str(data.get("Название (поставщик)") or "").strip()
        sku = str(data.get(f"К{n}: SKU") or "").strip()
        store_name = str(data.get(f"К{n}: Магазин") or "").strip()
        product_name = str(data.get(f"К{n}: Название") or "").strip()
        if not supplier or not supplier_name or not sku or not store_name:
            errors.append(f"Неполные данные: {supplier} / {supplier_name}"); continue
        store_id = stores_by_name.get(store_name)
        if store_id is None:
            errors.append(f"Магазин не найден: {store_name}"); continue
        match = db.query(ProductMatch).filter(
            ProductMatch.supplier == supplier,
            ProductMatch.supplier_name == supplier_name,
            ProductMatch.status == "pending",
        ).first()
        if not match:
            errors.append(f"Матч не найден: {supplier} / {supplier_name}"); continue
        match.sku = sku
        match.store_id = store_id
        match.product_name = product_name or None
        match.status = "auto_review"
        match.match_type = "manual_import"
        updated += 1
    db.commit()
    return {"updated": updated, "errors": errors}


@router.post("/import-sku-matches")
async def import_sku_matches(file: UploadFile, db: Session = Depends(get_db)):
    import io
    import openpyxl
    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active
    headers = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
    stores_by_name = {(s.display_name or s.name): s.id for s in db.query(Store).all()}
    stores_by_name.update({"ЯМ16": 1, "ЯМ21": 2})
    updated = 0
    errors = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        data = dict(zip(headers, row))
        choice_raw = data.get("Выбор (1-6)")
        if not choice_raw:
            continue
        try:
            n = int(str(choice_raw).strip())
        except ValueError:
            errors.append(f"Неверное значение Выбор: {choice_raw}"); continue
        if n < 1 or n > 6:
            errors.append(f"Выбор должен быть 1-6, получено: {n}"); continue
        store_name = str(data.get("Магазин") or "").strip()
        sku = str(data.get("SKU") or "").strip()
        supplier = str(data.get(f"К{n}: Поставщик") or "").strip()
        supplier_name_val = str(data.get(f"К{n}: Название") or "").strip()
        supplier_price_raw = data.get(f"К{n}: Цена закупки")
        if not store_name or not sku or not supplier or not supplier_name_val:
            errors.append(f"Неполные данные: {store_name} / {sku}"); continue
        store_id = stores_by_name.get(store_name)
        if store_id is None:
            errors.append(f"Магазин не найден: {store_name}"); continue
        try:
            supplier_price = float(supplier_price_raw) if supplier_price_raw else None
        except (ValueError, TypeError):
            supplier_price = None
        match = db.query(ProductMatch).filter(
            ProductMatch.supplier == supplier,
            ProductMatch.supplier_name == supplier_name_val,
            ProductMatch.status == "pending",
        ).first()
        product = db.query(Product).filter(Product.sku == sku, Product.store_id == store_id).first()
        if match:
            match.sku = sku
            match.store_id = store_id
            match.product_name = product.name if product else sku
            match.status = "auto_review"
            match.match_type = "manual_import"
            if supplier_price is not None:
                match.supplier_price = supplier_price
        else:
            sp = db.query(SupplierPrice).filter(
                SupplierPrice.supplier == supplier, SupplierPrice.name == supplier_name_val,
            ).first()
            db.add(ProductMatch(
                supplier=supplier, supplier_name=supplier_name_val,
                supplier_normalized=normalize_name(supplier_name_val),
                supplier_price=supplier_price or (sp.price if sp else None),
                sku=sku, store_id=store_id,
                product_name=product.name if product else sku,
                status="auto_review", match_type="manual_import",
            ))
        updated += 1
    db.commit()
    return {"updated": updated, "errors": errors}


@router.post("/backfill-embeddings")
def backfill_embeddings(db: Session = Depends(get_db)):
    if not embedding_service.is_available():
        raise HTTPException(status_code=503, detail="Ollama недоступна. Запустите: ollama serve")
    products_done = 0
    supplier_done = 0
    for p in db.query(Product).filter(Product.name_embedding.is_(None)).all():
        text = p.normalized_name or p.name or ""
        if text:
            emb = embedding_service.embed(text)
            if emb:
                p.name_embedding = json.dumps(emb)
                products_done += 1
    db.commit()
    for sp in db.query(SupplierPrice).filter(SupplierPrice.name_embedding.is_(None)).all():
        if sp.normalized_name:
            emb = embedding_service.embed(sp.normalized_name)
            if emb:
                sp.name_embedding = json.dumps(emb)
                supplier_done += 1
    db.commit()
    return {"products_embedded": products_done, "supplier_prices_embedded": supplier_done}
