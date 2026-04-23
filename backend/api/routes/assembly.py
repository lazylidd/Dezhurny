import logging
from typing import List, Optional
from urllib.parse import quote

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import Response as FResponse, StreamingResponse
from pydantic import BaseModel
from sqlalchemy.orm import Session

from api.deps import get_db
from models.product import Product
from models.store import Store
from services.yam_client import get_client_for_store
from utils.settings import get_setting

logger = logging.getLogger(__name__)
router = APIRouter(tags=["assembly"])


def _cd(filename: str) -> str:
    ascii_name = filename.encode("ascii", "ignore").decode()
    return f"attachment; filename=\"{ascii_name}\"; filename*=UTF-8''{quote(filename, safe='')}"


def _assembly_process_order(order, status_override, store, campaign_id, db):
    from models.product_match import ProductMatch as PM
    order_id = order.get("id")
    substatus = order.get("substatus", "")
    if status_override:
        status = status_override
    else:
        status = "READY_TO_SHIP" if substatus == "READY_TO_SHIP" else order.get("status", "")
    items = order.get("items") or []
    result_items = []
    for item in items:
        sku = item.get("offerId") or item.get("sku") or ""
        offer_name = item.get("offerName") or item.get("offer", {}).get("name") or ""
        count = int(item.get("count") or 1)
        buyer_price = float(item.get("buyerPrice") or item.get("price") or 0)
        subsidies = item.get("subsidies") or []
        subsidy_per_unit = sum(float(s.get("amount", 0)) for s in subsidies)
        sale_price = buyer_price + subsidy_per_unit
        total_buyer = round(sale_price * count, 2)
        match = db.query(PM).filter(
            PM.sku == sku, PM.store_id == store.id, PM.status == "confirmed",
        ).order_by(PM.supplier_price.asc()).first() if sku else None
        supplier_price = float(match.supplier_price) if match and match.supplier_price else None
        total_supplier = round(supplier_price * count, 2) if supplier_price else None
        product = db.query(Product).filter(Product.store_id == store.id, Product.sku == sku).first() if sku else None
        tax_rate = float(store.tax_rate or 0.06)
        if product and product.ym_variable_rate is not None:
            ym_variable_rate = float(product.ym_variable_rate)
            ym_fixed_fee = float(product.ym_fixed_fee or 0)
        else:
            ym_variable_rate = float(product.commission or 0) / 100 if product and product.commission else 0.0
            ym_fixed_fee = 0.0
        ym_var_amount = round(sale_price * ym_variable_rate * count, 2)
        ym_fix_amount = round(ym_fixed_fee * count, 2)
        tax_fees = round(buyer_price * tax_rate * count, 2)
        fees = ym_var_amount + ym_fix_amount + tax_fees
        fee_source = "tariffs" if (product and product.ym_variable_rate is not None) else "commission"
        fee_details = []
        if ym_var_amount:
            fee_details.append({"type": "FEE", "label": "Комиссия ЯМ", "amount": ym_var_amount, "pct": round(ym_variable_rate * 100, 2)})
        if ym_fix_amount:
            fee_details.append({"type": "FIXED", "label": "Доставка/фикс.", "amount": ym_fix_amount, "pct": round(ym_fix_amount / total_buyer * 100, 2) if total_buyer else 0})
        fee_details.append({"type": "TAX", "label": f"Налог (УСН {round(tax_rate*100)}%)", "amount": tax_fees, "pct": round(tax_rate * 100, 1)})
        if subsidy_per_unit:
            fee_details.append({"type": "SUBSIDY", "label": "Субсидия ЯМ", "amount": round(subsidy_per_unit * count, 2), "pct": round(subsidy_per_unit / sale_price * 100, 2) if sale_price else 0})
        profit = round(total_buyer - fees - total_supplier, 2) if total_supplier is not None else None
        roi = round(profit / total_supplier, 4) if (profit is not None and total_supplier) else None
        ros = round(profit / total_buyer, 4) if (profit is not None and total_buyer) else None
        result_items.append({
            "order_id": str(order_id),
            "status": status,
            "campaign_id": campaign_id,
            "sku": sku,
            "offer_name": offer_name,
            "count": count,
            "buyer_price": sale_price,
            "total_buyer": total_buyer,
            "supplier_price": supplier_price,
            "total_supplier": total_supplier,
            "fees": fees,
            "fee_details": fee_details,
            "fee_source": fee_source,
            "profit": profit,
            "roi": roi,
            "ros": ros,
        })
    return result_items


@router.get("/assembly")
def get_assembly(db: Session = Depends(get_db)):
    from datetime import datetime, timezone, timedelta

    msk_now = datetime.now(timezone.utc) + timedelta(hours=3)
    msk_hhmm = msk_now.strftime("%H:%M")
    today_ddmmyyyy = msk_now.strftime("%d-%m-%Y")
    cutoff_time = get_setting("order_cutoff_time") or "10:00"
    before_cutoff = msk_hhmm < cutoff_time

    def _ship_date(order: dict) -> str | None:
        delivery = order.get("delivery") or {}
        shipments = delivery.get("shipments") or []
        return shipments[0].get("shipmentDate") if shipments else None

    result = []
    stores = db.query(Store).order_by(Store.name).all()
    for store in stores:
        try:
            client, business_id, campaign_ids = get_client_for_store(store.name, db)
        except Exception:
            continue
        store_orders = []
        for campaign_id in campaign_ids:
            try:
                orders = client.get_orders_for_assembly(campaign_id)
            except Exception as e:
                logger.warning("[assembly] store=%s campaign=%s ошибка: %s", store.name, campaign_id, e)
                orders = []
            for order in orders:
                if _ship_date(order) == today_ddmmyyyy:
                    store_orders.extend(_assembly_process_order(order, None, store, campaign_id, db))

            try:
                delivery_orders = client.get_delivery_orders_today(campaign_id)
            except Exception as e:
                logger.warning("[assembly] DELIVERY store=%s campaign=%s ошибка: %s", store.name, campaign_id, e)
                delivery_orders = []
            for order in delivery_orders:
                if _ship_date(order) == today_ddmmyyyy:
                    store_orders.extend(_assembly_process_order(order, "DELIVERY", store, campaign_id, db))

        result.append({
            "store_id": store.id,
            "store_name": store.display_name or store.name,
            "campaign_ids": campaign_ids,
            "orders": store_orders,
        })
    return {
        "before_cutoff": before_cutoff,
        "cutoff_time": cutoff_time,
        "stores": result,
    }


@router.get("/assembly/purchase-list.pdf")
def download_purchase_list_pdf(db: Session = Depends(get_db)):
    import io
    import os
    from datetime import datetime, timezone, timedelta
    from fpdf import FPDF

    msk_now = datetime.now(timezone.utc) + timedelta(hours=3)
    today_ddmmyyyy = msk_now.strftime("%d-%m-%Y")
    today_str = msk_now.strftime("%d.%m.%Y")

    all_items: list = []
    stores_list = db.query(Store).all()
    for store in stores_list:
        try:
            client, _, campaign_ids = get_client_for_store(store.name, db=db)
        except Exception:
            continue
        for campaign_id in campaign_ids:
            try:
                orders = client.get_orders_for_assembly(campaign_id)
                for order in orders:
                    all_items.extend(_assembly_process_order(order, None, store, campaign_id, db))
            except Exception:
                pass
            try:
                delivery_orders = client.get_orders_for_assembly(campaign_id, status="DELIVERY")
                for order in delivery_orders:
                    all_items.extend(_assembly_process_order(order, "DELIVERY", store, campaign_id, db))
            except Exception:
                pass

    agg: dict = {}
    for item in all_items:
        name = item.get("offer_name") or item.get("sku") or "—"
        if name not in agg:
            agg[name] = {"count": 0, "supplier_price": item.get("supplier_price"), "total": 0.0}
        agg[name]["count"] += item.get("count", 1)
        agg[name]["total"] += item.get("total_supplier") or 0.0
        if agg[name]["supplier_price"] is None and item.get("supplier_price") is not None:
            agg[name]["supplier_price"] = item["supplier_price"]

    rows = sorted(agg.items(), key=lambda x: x[0])
    grand_total = sum(r["total"] for _, r in rows)
    total_orders = len(set(i.get("order_id", "") for i in all_items))

    font_path = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), "data", "fonts", "DejaVuSans.ttf")

    pdf = FPDF()
    pdf.add_page()
    pdf.add_font("DejaVu", fname=font_path)
    pdf.add_font("DejaVu", style="B", fname=font_path)

    pdf.set_font("DejaVu", style="B", size=16)
    pdf.cell(0, 10, "Закупочный лист", ln=True)
    pdf.set_font("DejaVu", size=10)
    pdf.set_text_color(100, 100, 100)
    pdf.cell(0, 7, f"Дата: {today_str}  ·  Позиций: {len(rows)}  ·  Заказов: {total_orders}", ln=True)
    pdf.set_text_color(0, 0, 0)
    pdf.ln(4)

    col_w = [10, 105, 20, 30, 30]
    headers = ["#", "Наименование", "Кол.", "Цена закупа", "Сумма"]
    pdf.set_font("DejaVu", style="B", size=9)
    pdf.set_fill_color(240, 242, 245)
    for w, h in zip(col_w, headers):
        align = "R" if h in ("Кол.", "Цена закупа", "Сумма") else "L"
        pdf.cell(w, 8, h, border=1, fill=True, align=align)
    pdf.ln()

    pdf.set_font("DejaVu", size=9)
    line_h = 6
    for i, (name, r) in enumerate(rows, 1):
        fill = i % 2 == 0
        pdf.set_fill_color(249, 250, 251) if fill else pdf.set_fill_color(255, 255, 255)
        price_str = f"{int(r['supplier_price']):,} ₽".replace(",", " ") if r["supplier_price"] else "—"
        total_str = f"{int(r['total']):,} ₽".replace(",", " ") if r["total"] > 0 else "—"
        x0 = pdf.l_margin
        y0 = pdf.get_y()
        pdf.set_xy(x0 + col_w[0], y0)
        pdf.multi_cell(col_w[1], line_h, name, border=1, fill=fill, align="L")
        row_bottom = pdf.get_y()
        row_h = max(row_bottom - y0, line_h)
        pdf.set_xy(x0, y0)
        pdf.cell(col_w[0], row_h, str(i), border=1, fill=fill, align="L")
        pdf.set_xy(x0 + col_w[0] + col_w[1], y0)
        pdf.cell(col_w[2], row_h, str(r["count"]), border=1, fill=fill, align="R")
        pdf.set_xy(x0 + col_w[0] + col_w[1] + col_w[2], y0)
        pdf.cell(col_w[3], row_h, price_str, border=1, fill=fill, align="R")
        pdf.set_xy(x0 + col_w[0] + col_w[1] + col_w[2] + col_w[3], y0)
        pdf.cell(col_w[4], row_h, total_str, border=1, fill=fill, align="R")
        pdf.set_xy(x0, row_bottom)

    pdf.set_font("DejaVu", style="B", size=9)
    pdf.set_fill_color(229, 231, 235)
    total_str = f"{grand_total:,.0f} ₽".replace(",", " ")
    pdf.cell(col_w[0] + col_w[1] + col_w[2] + col_w[3], 8, "ИТОГО", border=1, fill=True, align="R")
    pdf.cell(col_w[4], 8, total_str, border=1, fill=True, align="R")
    pdf.ln()

    buf = io.BytesIO(bytes(pdf.output()))
    fname = f"purchase_list_{today_str}.pdf"
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=\"{fname}\""},
    )


@router.get("/stores/{store_id}/assembly/shipments")
def get_assembly_shipments(store_id: int, db: Session = Depends(get_db)):
    store = db.query(Store).filter(Store.id == store_id).first()
    if not store:
        raise HTTPException(status_code=404)
    try:
        client, _, campaign_ids = get_client_for_store(store.name, db)
    except Exception as e:
        raise HTTPException(status_code=502, detail=str(e))
    all_shipments = []
    for campaign_id in campaign_ids:
        shipments = client.get_shipments(campaign_id)
        for s in shipments:
            all_shipments.append({**s, "campaign_id": campaign_id})
    return all_shipments


@router.post("/stores/{store_id}/assembly/ready")
def mark_orders_ready(store_id: int, campaign_id: int, order_ids: str, db: Session = Depends(get_db)):
    store = db.query(Store).filter(Store.id == store_id).first()
    if not store:
        raise HTTPException(status_code=404)
    try:
        client, _, _ = get_client_for_store(store.name, db)
    except Exception as e:
        raise HTTPException(status_code=502, detail=str(e))
    ids = [x.strip() for x in order_ids.split(",") if x.strip()]
    if not ids:
        raise HTTPException(status_code=400, detail="Нет order_ids")
    results = []
    for oid in ids:
        try:
            client.set_order_ready(campaign_id, oid)
            results.append({"order_id": oid, "ok": True})
        except Exception as e:
            results.append({"order_id": oid, "ok": False, "error": str(e)})
    return {"results": results}


@router.get("/stores/{store_id}/assembly/act")
def download_assembly_act(store_id: int, campaign_id: int, shipment_id, db: Session = Depends(get_db)):
    store = db.query(Store).filter(Store.id == store_id).first()
    if not store:
        raise HTTPException(status_code=404)
    try:
        client, _, _ = get_client_for_store(store.name, db)
    except Exception as e:
        raise HTTPException(status_code=502, detail=str(e))
    try:
        pdf = client.get_shipment_act(campaign_id, shipment_id)
    except Exception as e:
        raise HTTPException(status_code=502, detail=str(e))
    return FResponse(
        content=pdf,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=act_{campaign_id}_{shipment_id}.pdf"},
    )


class CampaignLabelsRequest(BaseModel):
    campaign_id: int
    order_ids: List[str]


class CampaignDocRequest(BaseModel):
    campaign_id: int
    order_ids: List[str]
    shipment_id: Optional[int] = None


class StoreDocRequest(BaseModel):
    store_id: int
    store_name: str
    campaigns: List[CampaignDocRequest]


@router.post("/stores/{store_id}/assembly/all-labels")
def download_all_labels(store_id: int, campaigns: List[CampaignLabelsRequest], db: Session = Depends(get_db)):
    import io as _io
    store = db.query(Store).filter(Store.id == store_id).first()
    if not store:
        raise HTTPException(status_code=404)
    try:
        client, _, _ = get_client_for_store(store.name, db)
    except Exception as e:
        raise HTTPException(status_code=502, detail=str(e))

    all_pdfs: list = []
    for camp in campaigns:
        ids = [int(x.strip()) for x in camp.order_ids if x.strip()]
        if not ids:
            continue
        try:
            pdf = client.get_labels_pdf(camp.campaign_id, ids)
            if pdf:
                all_pdfs.append(pdf)
        except Exception as e:
            raise HTTPException(status_code=502, detail=f"Ярлыки campaign {camp.campaign_id}: {e}")

    if not all_pdfs:
        raise HTTPException(status_code=400, detail="Нет ярлыков")

    from services.yam_client import _merge_pdfs
    merged = _merge_pdfs(all_pdfs)

    return FResponse(
        content=merged,
        media_type="application/pdf",
        headers={"Content-Disposition": _cd(f"ярлыки_{store.name}.pdf")},
    )


@router.post("/assembly/all-documents")
def download_all_documents(stores: List[StoreDocRequest], db: Session = Depends(get_db)):
    import io as _io
    import zipfile
    from concurrent.futures import ThreadPoolExecutor
    from datetime import date

    today = date.today().strftime("%d.%m.%Y")
    buf = _io.BytesIO()

    store_clients = {}
    for store_req in stores:
        try:
            store_obj = db.query(Store).filter(Store.id == store_req.store_id).first()
            if not store_obj:
                continue
            client, _, _ = get_client_for_store(store_obj.name, db)
            store_clients[store_req.store_id] = (client, store_obj.name)
        except Exception:
            continue

    def fetch_labels(client, campaign_id, order_ids, zip_path):
        try:
            pdf = client.get_labels_pdf(campaign_id, order_ids)
            return zip_path, pdf
        except Exception:
            return zip_path, None

    def fetch_shipment_list(client, campaign_id, shipment_id, zip_path, order_ids=None):
        try:
            pdf = client.get_shipment_list_pdf(campaign_id, shipment_id=shipment_id, order_ids=order_ids)
            return zip_path, pdf
        except Exception:
            return zip_path, None

    futures = []
    with ThreadPoolExecutor(max_workers=8) as executor:
        for store_req in stores:
            if store_req.store_id not in store_clients:
                continue
            client, store_name = store_clients[store_req.store_id]
            folder = f"отгрузка {today}/{store_name}"
            for camp in store_req.campaigns:
                if camp.order_ids:
                    futures.append(executor.submit(
                        fetch_labels, client, camp.campaign_id, camp.order_ids,
                        f"{folder}/ярлыки_{store_name}.pdf"
                    ))
                if camp.order_ids:
                    futures.append(executor.submit(
                        fetch_shipment_list, client, camp.campaign_id, camp.shipment_id,
                        f"{folder}/лист_сборки_{store_name}.pdf", camp.order_ids
                    ))

        results = [f.result() for f in futures]

    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for zip_path, pdf in results:
            if pdf:
                zf.writestr(zip_path, pdf)

    buf.seek(0)
    fname = f"сборка_{today}.zip"
    return FResponse(
        content=buf.read(),
        media_type="application/zip",
        headers={"Content-Disposition": _cd(fname)},
    )


@router.get("/stores/{store_id}/assembly/shipment-list")
def download_shipment_list(
    store_id: int, campaign_id: int,
    order_ids: Optional[str] = None,
    shipment_id: Optional[int] = None,
    db: Session = Depends(get_db),
):
    store = db.query(Store).filter(Store.id == store_id).first()
    if not store:
        raise HTTPException(status_code=404)
    try:
        client, _, _ = get_client_for_store(store.name, db)
    except Exception as e:
        raise HTTPException(status_code=502, detail=str(e))
    ids = [x.strip() for x in order_ids.split(",") if x.strip()] if order_ids else None
    try:
        pdf = client.get_shipment_list_pdf(campaign_id, shipment_id=shipment_id, order_ids=ids)
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"Ошибка генерации листа сборки: {e}")
    return FResponse(
        content=pdf,
        media_type="application/pdf",
        headers={"Content-Disposition": _cd(f"лист_сборки_{store.name}.pdf")},
    )


@router.get("/stores/{store_id}/assembly/labels")
def download_assembly_labels(store_id: int, campaign_id: int, order_ids: str, db: Session = Depends(get_db)):
    store = db.query(Store).filter(Store.id == store_id).first()
    if not store:
        raise HTTPException(status_code=404)
    try:
        client, _, _ = get_client_for_store(store.name, db)
    except Exception as e:
        raise HTTPException(status_code=502, detail=str(e))
    ids = [int(x.strip()) for x in order_ids.split(",") if x.strip()]
    if not ids:
        raise HTTPException(status_code=400, detail="Нет order_ids")
    try:
        pdf = client.get_labels_pdf(campaign_id, ids)
    except Exception as e:
        raise HTTPException(status_code=502, detail=str(e))
    return FResponse(
        content=pdf,
        media_type="application/pdf",
        headers={"Content-Disposition": _cd(f"ярлыки_{store.name}.pdf")},
    )


@router.get("/stores/{store_id}/assembly/sheet")
def download_assembly_sheet(store_id: int, campaign_id: int, db: Session = Depends(get_db)):
    import io as _io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from datetime import date

    store = db.query(Store).filter(Store.id == store_id).first()
    if not store:
        raise HTTPException(status_code=404)
    try:
        client, _, campaign_ids = get_client_for_store(store.name, db)
    except Exception as e:
        raise HTTPException(status_code=502, detail=str(e))
    if campaign_id not in campaign_ids:
        raise HTTPException(status_code=400, detail="campaign_id не принадлежит магазину")
    try:
        orders = client.get_orders_for_assembly(campaign_id)
    except Exception as e:
        raise HTTPException(status_code=502, detail=str(e))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Лист сборки"
    header_fill = PatternFill(start_color="374151", end_color="374151", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    headers = ["№", "Заказ", "Статус", "SKU", "Товар", "Кол-во", "Цена покупателя", "Итого"]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    row_num = 1
    for order in orders:
        order_id = order.get("id")
        status = order.get("status", "")
        for item in (order.get("items") or []):
            row_num += 1
            sku = item.get("offerId") or ""
            name = item.get("offerName") or ""
            count = int(item.get("count") or 1)
            buyer_price = float(item.get("buyerPrice") or item.get("price") or 0)
            subsidies = item.get("subsidies") or []
            subsidy_per_unit = sum(float(s.get("amount", 0)) for s in subsidies)
            price = buyer_price + subsidy_per_unit
            ws.append([row_num - 1, str(order_id), status, sku, name, count, price, round(price * count, 2)])
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)
    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"assembly_{store.name}_{campaign_id}_{date.today().isoformat()}.xlsx"
    return FResponse(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": _cd(fname)},
    )
