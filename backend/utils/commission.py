from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook


class CommissionError(RuntimeError):
    pass


def normalize_cat_part(value: Any) -> str:
    if value is None:
        return ""
    s = str(value).strip()
    s = " ".join(s.split())
    return s.casefold()


def build_path(parts: List[Any]) -> str:
    clean = [str(x).strip() for x in parts if x not in (None, "")]
    return " > ".join(clean)


def build_norm_path(parts: List[Any]) -> str:
    clean = [normalize_cat_part(x) for x in parts if x not in (None, "")]
    clean = [x for x in clean if x]
    return " > ".join(clean)


def parse_commission_value(value: Any) -> Any:
    if value is None or value == "":
        return ""
    try:
        return round(float(value) * 100, 4)
    except Exception:
        return value


def load_commission_paths_from_excel(xlsx_path: str, sheet_name: str = "FBS с 1.02.2026") -> Dict[str, Any]:
    wb = load_workbook(xlsx_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise CommissionError(f"Лист '{sheet_name}' не найден в файле {xlsx_path}")

    ws = wb[sheet_name]
    headers = [ws.cell(row=1, column=i).value for i in range(1, ws.max_column + 1)]
    header_to_idx = {str(v).strip(): i + 1 for i, v in enumerate(headers) if v is not None}

    level_cols = []
    for n in range(1, 8):
        key = f"Категория (Уровень {n})"
        if key not in header_to_idx:
            raise CommissionError(f"В Excel нет колонки '{key}'")
        level_cols.append(header_to_idx[key])

    tariff_col_name = None
    for name in header_to_idx:
        if str(name).startswith("Тариф с "):
            tariff_col_name = name
            break

    if not tariff_col_name:
        raise CommissionError("В Excel не найдена колонка вида 'Тариф с ...'")

    tariff_col = header_to_idx[tariff_col_name]
    path_to_commission: Dict[str, Any] = {}

    for row in range(2, ws.max_row + 1):
        parts = [ws.cell(row=row, column=col).value for col in level_cols]
        raw_commission = ws.cell(row=row, column=tariff_col).value

        norm_path = build_norm_path(parts)
        if not norm_path:
            continue

        parsed_commission = parse_commission_value(raw_commission)

        if norm_path in path_to_commission and path_to_commission[norm_path] != parsed_commission:
            raise CommissionError(
                f"Дублирующийся путь категории с разными комиссиями: {norm_path}"
            )

        path_to_commission[norm_path] = parsed_commission

    return path_to_commission


def flatten_category_tree(node: Dict[str, Any], parent_id: Optional[int], out: Dict[int, Dict[str, Any]]) -> None:
    category = node.get("category") or node
    cid = category.get("id")
    name = category.get("name")

    if cid is not None:
        out[int(cid)] = {"id": int(cid), "name": name or "", "parent_id": parent_id}

    children = category.get("children") or node.get("children") or []
    for child in children:
        flatten_category_tree(child, int(cid) if cid is not None else None, out)


def build_category_maps(tree_response: Dict[str, Any]) -> Tuple[Dict[int, Dict[str, Any]], Dict[int, str], Dict[int, str]]:
    result = tree_response.get("result", {}) or {}
    roots = result.get("children") or result.get("categories") or []

    categories_by_id: Dict[int, Dict[str, Any]] = {}
    for root in roots:
        flatten_category_tree(root, None, categories_by_id)

    full_path_by_id: Dict[int, str] = {}
    norm_full_path_by_id: Dict[int, str] = {}

    for cid in categories_by_id:
        parts = []
        cur = cid
        while cur is not None and cur in categories_by_id:
            parts.append(categories_by_id[cur]["name"])
            cur = categories_by_id[cur]["parent_id"]
        parts.reverse()
        full_path_by_id[cid] = build_path(parts)
        norm_full_path_by_id[cid] = build_norm_path(parts)

    return categories_by_id, full_path_by_id, norm_full_path_by_id


def find_commission_for_category(
    category_id: Any,
    categories_by_id: Dict[int, Dict[str, Any]],
    norm_full_path_by_id: Dict[int, str],
    commission_by_path: Dict[str, Any],
) -> Tuple[Any, str, str]:
    if category_id in (None, ""):
        return "", "", ""

    try:
        cid = int(category_id)
    except Exception:
        return "", "", ""

    full_path = norm_full_path_by_id.get(cid, "")
    if not full_path:
        return "", "", ""

    parts = full_path.split(" > ")
    for i in range(len(parts), 0, -1):
        sub_path = " > ".join(parts[:i])
        if sub_path in commission_by_path:
            match_type = "exact" if i == len(parts) else "parent_fallback"
            return commission_by_path[sub_path], sub_path, match_type

    return "", "", ""
